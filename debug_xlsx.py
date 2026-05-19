import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\crush\Downloads\Telegram Desktop\Прайс Лаборатория.xlsx')
ws = wb.active
print(f"Sheet: {ws.title}, Rows: {ws.max_row}, Cols: {ws.max_column}")

for i, row in enumerate(ws.iter_rows(min_col=1, max_col=min(ws.max_column, 2), values_only=True), 1):
    print(f"Row {i}: colA={repr(row[0])}, colB={repr(row[1]) if len(row)>1 else 'N/A'}")

wb.close()
